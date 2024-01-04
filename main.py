import asyncio
import os

import pytz
import aiohttp
import concurrent.futures

from capmonstercloudclient.requests import HcaptchaProxylessRequest
from dotenv import load_dotenv


from datetime import datetime
from openpyxl import load_workbook
from twocaptcha import TwoCaptcha

from bs4 import BeautifulSoup

from capmonstercloudclient import CapMonsterClient, ClientOptions

# Load env variables
load_dotenv()
number_of_accounts = int(os.getenv('ACCOUNTS_COUNT'))
USER_ID = os.getenv('USER_ID')
CAPMONSTER_KEY = os.getenv('CAPMONSTER_KEY')

almaty_tz = pytz.timezone('Asia/Almaty')

TWOCAPTHCA = "twocaptcha"
REHALKA = "rehalka"
CAPMONSTER = "capmonster"

REHALKA_IP = "188.124.36.201:3005"
CAPTCHA_URL = "https://evergem.io/claim?redirect_to=/game"
CAPTCHA_KEY = "d1add268-b915-46c1-afd3-960faba20822"

TELEGRAM_BOT_TOKEN = "6933846503:AAEL6QXu_4a9yEyiX-yZua2MyeEMyHAn0IA"

client_options = ClientOptions(api_key=CAPMONSTER_KEY)
cap_monster_client = CapMonsterClient(options=client_options)

workbook = load_workbook(filename='info.xlsx')
sheet = workbook.active
data = []

min_row = 2
max_row = min_row + number_of_accounts - 1
for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
    data.append({
        'name': row[0],
        'proxy': row[1],
        'ua': row[2],
        'item_id': row[7],
        'token': row[8],
        'cookie': row[9],
        'rehalka_key': row[10]
    })


async def send_telegram_message(bot_token, chat_id, text):
    url = f'https://api.telegram.org/bot{bot_token}/sendMessage'
    params = {'chat_id': chat_id, 'text': text}
    async with aiohttp.ClientSession() as tg_session:
        async with tg_session.post(url, params=params) as response:
            response.raise_for_status()


async def captchaSolver(method, rehalka_key):
    if method == TWOCAPTHCA:
        loop = asyncio.get_running_loop()
        with concurrent.futures.ThreadPoolExecutor() as pool:
            two_result = await loop.run_in_executor(pool, lambda: TwoCaptcha("5bcf672e872e4ed67e1b0a1627eece17",
                                                                         defaultTimeout=180).hcaptcha(
                sitekey=CAPTCHA_URL,
                url=CAPTCHA_KEY,
            ))
            return two_result.get('code').replace(" ", "")
    elif method == CAPMONSTER:
        recaptcha2request = HcaptchaProxylessRequest(
            websiteUrl=CAPTCHA_URL,
            websiteKey=CAPTCHA_KEY)
        cap_result = await cap_monster_client.solve_captcha(recaptcha2request)
        return cap_result['gRecaptchaResponse']
    elif method == REHALKA:
        async with aiohttp.ClientSession() as rh_session:
            url = f"http://{REHALKA_IP}/in.php?userkey={rehalka_key}&host={CAPTCHA_URL}&sitekey={CAPTCHA_KEY}"
            captcha_id = None
            while captcha_id is None:
                await asyncio.sleep(1)

                async with rh_session.get(url) as in_response:
                    in_response.raise_for_status()
                    captcha_info = await in_response.text()
                if captcha_info in ["ERROR_KEY_DOES_NOT_EXIST", "ERROR_ZERO_BALANCE",
                                     "ERROR_CAPTCHA_UNSOLVABLE", "ERROR_WRONG_SAITKEY", "ERROR_WRONG_CAPTCHA_ID"]:
                    raise ValueError(f"Getting result: {captcha_info}")
                elif captcha_info in ["ERROR_NO_SLOT_AVAILABLE", "CAPCHA_NOT_READY"]:
                    continue
                elif captcha_info.startswith("OK"):
                    captcha_id = captcha_info.split('|')[1]

            while True:
                await asyncio.sleep(5)

                async with aiohttp.ClientSession() as rh_res_session:
                    url = f"http://{REHALKA_IP}/res.php?userkey={rehalka_key}&id={captcha_id}"
                    async with rh_res_session.get(url) as res_response:
                        res_response.raise_for_status()
                        response_text = await res_response.text()

                    if response_text in ["ERROR_KEY_DOES_NOT_EXIST", "ERROR_ZERO_BALANCE",
                                  "ERROR_CAPTCHA_UNSOLVABLE", "ERROR_WRONG_SAITKEY", "ERROR_WRONG_CAPTCHA_ID"]:
                        raise ValueError(f"Getting result: {response_text}")
                    elif response_text in ["ERROR_NO_SLOT_AVAILABLE", "CAPCHA_NOT_READY"]:
                        continue
                    elif response_text.startswith("OK"):
                        return response_text.split("|")[1]

    return None


async def work(account):
    async with aiohttp.ClientSession() as session:

        claim_count = 1

        last_proxy_log_time = None

        last_withdraw_time = None
        last_withdraw_log_time = None

        name = account['name']
        proxy = account['proxy']
        user_agent = account['ua']
        item_id = account['item_id']
        token = account['token']
        cookie = account['cookie']
        rehalka_key = account['rehalka_key']
        # print(f"{name} данные загружены")

        while True:
            start_cycle_dt = datetime.now(almaty_tz)
            try:
                # current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                # print(f"{current_time}: {name} начинаю решать капчу")
                gh = await captchaSolver(CAPMONSTER, "")

                # current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                # print(f"{current_time}: {name} решил капчу")
                # print(gh)
            except Exception as e:
                current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                print(f"{current_time}: {name}: Ошибка в решении капчи: {e}")
            else:
                try:
                    # current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                    # print(f"{current_time}: {name} начинаю клейм")
                    url = 'https://evergem.io/claim'
                    params = {
                        'redirect_to': '/game'
                    }
                    payload = {
                        'item_id': f'{item_id}',
                        'h-captcha-response': gh,
                        'castle_request_token': token
                    }
                    headers = {
                        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Accept-Language': 'uk,ru-RU;q=0.9,ru;q=0.8,en-US;q=0.7,en;q=0.6,bg;q=0.5,es;q=0.4',
                        'Cache-Control': 'max-age=0',
                        'Content-Type': 'application/x-www-form-urlencoded',
                        'Cookie': f'{cookie}',
                        'Origin': 'null',
                        'Sec-Ch-Ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
                        'Sec-Ch-Ua-Mobile': '?0',
                        'Sec-Ch-Ua-Platform': '"Windows"',
                        'Sec-Fetch-Dest': 'document',
                        'Sec-Fetch-Mode': 'navigate',
                        'Sec-Fetch-Site': 'same-origin',
                        'Sec-Fetch-User': '?1',
                        'Upgrade-Insecure-Requests': '1',
                        'User-Agent': f'{user_agent}'
                    }

                    proxy_url = None
                    if proxy.lower() != 'no':
                        proxy_creds = proxy.split(':')
                        proxy_url = f"http://{proxy_creds[2]}:{proxy_creds[3]}@{proxy_creds[0]}:{proxy_creds[1]}"
                    # current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                    # print(f"{current_time}: {name} перед запросом")
                    if proxy_url:
                        async with session.post(url,
                                                headers=headers,
                                                params=params,
                                                data=payload,
                                                proxy=proxy_url) as response:
                            # current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                            # print(f"{current_time}: {name} внутри запроса с прокси")
                            response.raise_for_status()
                            text = await response.text()
                    else:
                        async with session.post(url,
                                                headers=headers,
                                                params=params,
                                                data=payload) as response:
                            # current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                            # print(f"{current_time}: {name} внутри запроса без прокси")
                            response.raise_for_status()
                            text = await response.text()
                except Exception as e:
                    current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                    print(f"{current_time}: {name}: Ошибка во время клейма: {e}")

                    if response.status == 403:
                        current_time_raw = datetime.now(almaty_tz)
                        if last_proxy_log_time is None or (
                                current_time_raw - last_proxy_log_time).total_seconds() > 3600:
                            log_msg = f"Account: {name}. Proxy problem! proxy: {proxy_url}"
                            await send_telegram_message(TELEGRAM_BOT_TOKEN, USER_ID, log_msg)
                            last_proxy_log_time = current_time_raw
                else:
                    soup = BeautifulSoup(text, 'html.parser')
                    balance_info_container = soup.find('div', class_='balance-info')
                    top_wrap_container = soup.find('div', class_='top-wrap')

                    balance_info_p_tags = balance_info_container.find_all('p')
                    balance_hpe = next((tag.text.strip() for tag in balance_info_p_tags if "HPE" in tag.text), None)
                    balance_usd = next((tag.text.strip() for tag in balance_info_p_tags if "$" in tag.text), None)

                    rate_wrap_container = top_wrap_container.find('div', class_='rate-wrap')
                    rate_hpe_day = next((tag.text.strip() for tag in rate_wrap_container.find_all('span') if "HPE/day" in tag.text),None)

                    current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                    print(f"{current_time}: {name}: Удачный клейм: {claim_count}, Баланс: {balance_hpe} ({balance_usd}) | {rate_hpe_day}")
                    claim_count += 1

                    balance_hpe_float = float(balance_hpe.split(' ')[0])
                    rate_hpe_float_day = float(rate_hpe_day.split(' ')[0])

                    current_time_raw = datetime.now(almaty_tz)
                    if balance_hpe_float >= rate_hpe_float_day/4:
                        try:
                            # current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                            # print(f"{current_time}: {name}: Вывожу HPE: {balance_hpe_float}")

                            if last_withdraw_time is not None and (
                                    current_time_raw - last_withdraw_time).total_seconds() <= 600:
                                if last_withdraw_log_time is None or (
                                        current_time_raw - last_withdraw_log_time).total_seconds() > 3600:
                                    log_msg = f"Account: {name}. Withdraw problem! Balance: {balance_usd}"
                                    await send_telegram_message(TELEGRAM_BOT_TOKEN, USER_ID, log_msg)
                                    last_withdraw_log_time = current_time_raw
                                raise ValueError("Уже выводил этот час")

                            url = 'https://evergem.io/withdraw'
                            params = {
                                'redirect_to': '/game'
                            }
                            payload = {
                                'amount': balance_hpe_float
                            }
                            headers = {
                                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                                'Accept-Encoding': 'gzip, deflate, br',
                                'Accept-Language': 'uk,ru-RU;q=0.9,ru;q=0.8,en-US;q=0.7,en;q=0.6,bg;q=0.5,es;q=0.4',
                                'Cache-Control': 'max-age=0',
                                'Content-Type': 'application/x-www-form-urlencoded',
                                'Cookie': f'{cookie}',
                                'Origin': 'null',
                                'Sec-Ch-Ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
                                'Sec-Ch-Ua-Mobile': '?0',
                                'Sec-Ch-Ua-Platform': '"Windows"',
                                'Sec-Fetch-Dest': 'document',
                                'Sec-Fetch-Mode': 'navigate',
                                'Sec-Fetch-Site': 'same-origin',
                                'Sec-Fetch-User': '?1',
                                'Upgrade-Insecure-Requests': '1',
                                'User-Agent': f'{user_agent}'
                            }

                            proxy_url = None
                            if proxy.lower() != 'no':
                                proxy_creds = proxy.split(':')
                                proxy_url = f"http://{proxy_creds[2]}:{proxy_creds[3]}@{proxy_creds[0]}:{proxy_creds[1]}"
                            # current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                            # print(f"{current_time}: {name} перед запросом")
                            async with aiohttp.ClientSession() as withdraw_session:
                                if proxy_url:
                                    async with withdraw_session.post(url,
                                                            headers=headers,
                                                            params=params,
                                                            data=payload,
                                                            proxy=proxy_url) as response:
                                        # current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                                        # print(f"{current_time}: {name} внутри запроса с прокси")
                                        response.raise_for_status()
                                else:
                                    async with withdraw_session.post(url,
                                                            headers=headers,
                                                            params=params,
                                                            data=payload) as response:
                                        # current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                                        # print(f"{current_time}: {name} внутри запроса без прокси")
                                        response.raise_for_status()
                        except Exception as e:
                            current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
                            print(f"{current_time}: {name}: Ошибка во время вывода токенов: {e}")
                        else:
                            current_time_raw = datetime.now(almaty_tz)
                            current_time = current_time_raw.strftime("%Y-%m-%d %H:%M:%S")
                            print(f"{current_time}: {name}: Удачный вывод: {balance_hpe_float} HPE")
                            last_withdraw_time = current_time_raw

            # current_time = datetime.now(almaty_tz).strftime("%Y-%m-%d %H:%M:%S")
            # print(f"{current_time}: {name} засыпает")
            end_cycle_dt = datetime.now(almaty_tz)
            cycle_seconds = (end_cycle_dt - start_cycle_dt).total_seconds()
            if cycle_seconds < 300:
                await asyncio.sleep(300 - int(cycle_seconds))


async def main():
    tasks = []

    for account in data:
        task = asyncio.create_task(work(account))
        tasks.append(task)

    await asyncio.gather(*tasks)


asyncio.run(main())
